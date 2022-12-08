import requests
from bs4 import  BeautifulSoup
from openpyxl import Workbook, load_workbook
# Excel file : workbook - save, load
# sheet in file : worksheet - record contents

wb = Workbook()
ws = wb.active

# ws.cell(1,1,'value')
# ws['A3'] = 'value'
# add last line
# ws.append(['value1','value2','value3'])
ws.append(['회차','1','2','3','4','5','6','bonus'])

for count in range(0, 10):
    url = 'https://search.daum.net/search?w=tot&DA=LOT&rtmaxcoll=LOT&&q=' + str(966 - count) + '%ED%9A%8C%EC%B0%A8%20%EB%A1%9C%EB%98%90'
    res = requests.get(url)

    if res.status_code == requests.codes.ok:
        print('정상접속')
        html = BeautifulSoup(res.text, 'html.parser') #형태 변환 파싱
        balls = html.select('.ball')                # 전체 요소 찾기
        del(balls[6])
        balls = [int(ball.text) for ball in balls]
        balls.sort()
        ws.append([count] + balls)

        # html.select_one()                         # 맨 처음 하나만
    else:
        print('접속실패')
wb.save('lotto.xlsx')


from openpyxl import Workbook, load_workbook
wb = load_workbook('lotto.xlsx')
ws = wb.active

# ws['A1'].value

for index, row in enumerate(ws.rows):
    for col in row:
        if index == 0:
            print(f"{col.value:^5s}", end='')
        else:
            print(f"{col.value:<5d}", end='')
    print()


for r in ws.rows:
    for col in r:
        print(col.value)

